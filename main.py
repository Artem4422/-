
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import asyncio
import threading
import json
import os
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime

# Когда собрано PyInstaller-ом — файлы кладём рядом с .exe, не во временную папку
if getattr(sys, "frozen", False):
    _APP_DIR = os.path.dirname(sys.executable)
else:
    _APP_DIR = os.path.dirname(os.path.abspath(__file__))

CONFIG_FILE  = os.path.join(_APP_DIR, "config.json")
RESULTS_FILE = os.path.join(_APP_DIR, "results.txt")
SESSION_FILE = os.path.join(_APP_DIR, "tg_session")


# ── Helpers ─────────────────────────────────────────────────────────────────

def _entity_name(e):
    if e is None:
        return "Неизвестно"
    if t := getattr(e, 'title', None):
        return t
    first = getattr(e, 'first_name', '') or ''
    last  = getattr(e, 'last_name',  '') or ''
    name  = f"{first} {last}".strip()
    if u := getattr(e, 'username', None):
        name += f" (@{u})"
    return name or str(getattr(e, 'id', '?'))

def _chat_title(c):
    if c is None:
        return "Неизвестный чат"
    return (getattr(c, 'title', None)
            or getattr(c, 'first_name', None)
            or str(getattr(c, 'id', '?')))


def _normalize_username(value: str) -> str:
    value = (value or "").strip()
    if value.startswith("@"):
        value = value[1:]
    value = value.split()[0] if value else ""
    if not value or value.lower() in {"нет username", "-", "—"}:
        return ""
    return value


def _col_letters(cell_ref: str) -> str:
    m = re.match(r"([A-Z]+)", cell_ref or "")
    return m.group(1) if m else ""


def _col_index(col: str) -> int:
    idx = 0
    for ch in col:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def _index_to_col(idx: int) -> str:
    idx += 1
    out = ""
    while idx:
        idx, rem = divmod(idx - 1, 26)
        out = chr(rem + ord("A")) + out
    return out


def _read_xlsx_sheet(path: str) -> dict[tuple[int, int], str]:
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    shared: list[str] = []
    cells: dict[tuple[int, int], str] = {}

    with zipfile.ZipFile(path) as zf:
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall(".//m:si", ns):
                parts = [t.text or "" for t in si.findall(".//m:t", ns)]
                shared.append("".join(parts))

        sheet_name = next((n for n in zf.namelist() if n.startswith("xl/worksheets/sheet")), None)
        if not sheet_name:
            raise ValueError("В файле Excel не найден лист")

        sheet = ET.fromstring(zf.read(sheet_name))
        for row in sheet.findall(".//m:sheetData/m:row", ns):
            row_idx = int(row.get("r", "0")) - 1
            for cell in row.findall("m:c", ns):
                ref = cell.get("r", "")
                col = _col_letters(ref)
                if not col:
                    continue
                col_idx = _col_index(col)
                val_node = cell.find("m:v", ns)
                if val_node is None or val_node.text is None:
                    continue
                val = val_node.text
                if cell.get("t") == "s":
                    val = shared[int(val)]
                cells[(row_idx, col_idx)] = str(val).strip()
    return cells


def parse_leads_xlsx(path: str) -> list[dict]:
    """Парсит Excel в формате: строка 1 — «Ключ», строка 2 — категории, далее @username + описание."""
    cells = _read_xlsx_sheet(path)
    if not cells:
        return []

    header_row = 1
    categories: list[tuple[int, str]] = []
    max_col = max(c for _, c in cells)
    for col in range(max_col + 1):
        title = cells.get((header_row, col), "")
        if title:
            categories.append((col, title))

    if not categories:
        raise ValueError("Не найдена строка категорий (ожидается вторая строка файла)")

    leads: list[dict] = []
    seen: set[str] = set()
    max_row = max(r for r, _ in cells)

    for row in range(header_row + 1, max_row + 1):
        for col, category in categories:
            username_raw = cells.get((row, col), "")
            if not username_raw:
                continue
            username = _normalize_username(username_raw)
            if not username:
                continue
            key = username.lower()
            if key in seen:
                continue
            seen.add(key)

            description = cells.get((row, col + 1), "")
            leads.append({
                "username": username,
                "name": f"@{username}",
                "chat": category,
                "message": description[:300],
                "source": "import",
            })
    return leads


def write_leads_xlsx(path: str, groups: list[tuple[str, list[dict]]]) -> int:
    """Пишет Excel в формате программы: «Ключ» → категории → @username + описание."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ImportError("Установите openpyxl: pip install openpyxl") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(1, 1, "Ключ")

    total = 0
    for i, (category, members) in enumerate(groups):
        col = 2 + i * 3
        ws.cell(2, col, category)
        for row_idx, member in enumerate(members):
            excel_row = 3 + row_idx
            username = _normalize_username(member.get("username", ""))
            if not username:
                continue
            ws.cell(excel_row, col, f"@{username}")
            ws.cell(excel_row, col + 1, (member.get("message") or "")[:500])
            total += 1

    wb.save(path)
    return total


def _edit_copy(widget, is_text: bool, root: tk.Tk):
    try:
        if is_text:
            if widget.tag_ranges(tk.SEL):
                root.clipboard_clear()
                root.clipboard_append(widget.get(tk.SEL_FIRST, tk.SEL_LAST))
        elif widget.selection_present():
            root.clipboard_clear()
            root.clipboard_append(widget.selection_get())
    except tk.TclError:
        pass


def _edit_paste(widget, is_text: bool, root: tk.Tk):
    try:
        text = root.clipboard_get()
    except tk.TclError:
        return
    try:
        if is_text:
            if str(widget.cget("state")) == tk.DISABLED:
                return
            if widget.tag_ranges(tk.SEL):
                widget.delete(tk.SEL_FIRST, tk.SEL_LAST)
            widget.insert(tk.INSERT, text)
        else:
            if widget.selection_present():
                widget.delete(tk.SEL_FIRST, tk.SEL_LAST)
            widget.insert(tk.INSERT, text)
    except tk.TclError:
        pass


def _edit_cut(widget, is_text: bool, root: tk.Tk):
    _edit_copy(widget, is_text, root)
    try:
        if is_text:
            if str(widget.cget("state")) != tk.DISABLED and widget.tag_ranges(tk.SEL):
                widget.delete(tk.SEL_FIRST, tk.SEL_LAST)
        elif widget.selection_present():
            widget.delete(tk.SEL_FIRST, tk.SEL_LAST)
    except tk.TclError:
        pass


def _edit_select_all(widget, is_text: bool):
    try:
        if is_text:
            widget.tag_add(tk.SEL, "1.0", tk.END)
            widget.mark_set(tk.INSERT, tk.END)
            widget.see(tk.INSERT)
        else:
            widget.select_range(0, tk.END)
            widget.icursor(tk.END)
    except tk.TclError:
        pass


def _bind_edit_menu(root: tk.Tk, widget, *, is_text: bool = False, readonly: bool = False):
    """ПКМ и Cmd+C/V/X/A — для Mac и Windows."""
    menu = tk.Menu(widget, tearoff=0)
    if not readonly:
        menu.add_command(
            label="Вырезать",
            command=lambda: _edit_cut(widget, is_text, root),
        )
    menu.add_command(
        label="Копировать",
        command=lambda: _edit_copy(widget, is_text, root),
    )
    if not readonly:
        menu.add_command(
            label="Вставить",
            command=lambda: _edit_paste(widget, is_text, root),
        )
    menu.add_separator()
    menu.add_command(
        label="Выделить всё",
        command=lambda: _edit_select_all(widget, is_text),
    )

    def show_menu(event):
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()
        return "break"

    for seq in _context_menu_bindings():
        widget.bind(seq, show_menu, add=True)


def _context_menu_bindings() -> tuple[str, ...]:
    """На Mac без Control+клик — он мешает обычным нажатиям."""
    if sys.platform == "darwin":
        return ("<Button-2>",)
    return ("<Button-3>", "<Control-Button-1>", "<Button-2>")


# ── Application ──────────────────────────────────────────────────────────────

class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Telegram Парсер")
        self.root.geometry("980x780")
        self.root.minsize(760, 580)

        self.client          = None
        self.is_monitoring   = False
        self.is_scanning     = False
        self.is_inviting     = False
        self._invite_stop    = False
        self._keywords: list = []
        self.matched_users: list = []
        self.invite_users: list = []
        self._invite_chat_entities: dict = {}
        self._invite_label_to_id: dict = {}
        self._last_export_path = ""
        self.is_exporting = False
        self._monitor_chat_entity = None
        self._active_handler = None
        self._ui_queue: list = []
        self._ui_coalesce: dict = {}
        self._ui_coalesce_pending = False
        self._poll_scheduled = False
        self._pending_result_lines: list[str] = []

        # Dedicated asyncio loop in background thread
        self.loop = asyncio.new_event_loop()
        threading.Thread(target=self.loop.run_forever, daemon=True).start()

        self._check_telethon()
        self._build_ui()
        self._setup_shortcuts()
        self._load_config()
        self._schedule_poll()

    # ── Telethon check ────────────────────────────────────────────────────────

    def _check_telethon(self):
        try:
            import telethon  # noqa: F401
        except ImportError:
            messagebox.showerror(
                "Отсутствует зависимость",
                "Библиотека Telethon не установлена.\n\n"
                "Запустите install.bat или выполните:\n"
                "  pip install telethon"
            )

    # ── Keyboard shortcuts ────────────────────────────────────────────────────

    def _setup_shortcuts(self):
        """Ctrl/Cmd + C/V/X/A по keycode — Mac и Windows, любая раскладка."""
        def on_edit_key(event):
            w = event.widget
            if isinstance(w, (ttk.Button, tk.Button)):
                return
            kc = event.keycode
            is_text = isinstance(w, tk.Text)
            is_entry = isinstance(w, (tk.Entry, ttk.Entry))

            if kc == 67:  # C
                _edit_copy(w, is_text, self.root)
            elif kc == 86 and not (is_text and str(w.cget("state")) == tk.DISABLED):  # V
                _edit_paste(w, is_text, self.root)
            elif kc == 88 and not (is_text and str(w.cget("state")) == tk.DISABLED):  # X
                _edit_cut(w, is_text, self.root)
            elif kc == 65 and (is_text or is_entry):  # A
                _edit_select_all(w, is_text)

        for mod in ("Control", "Command"):
            self.root.bind_all(f"<{mod}-KeyPress>", on_edit_key)

    # ── UI construction ───────────────────────────────────────────────────────

    def _build_ui(self):
        nb = ttk.Notebook(self.root)
        nb.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        tab1 = ttk.Frame(nb); nb.add(tab1, text="  Подключение  ")
        tab2 = ttk.Frame(nb); nb.add(tab2, text="  Мониторинг  ")
        tab3 = ttk.Frame(nb); nb.add(tab3, text="  Рассылка  ")
        tab4 = ttk.Frame(nb); nb.add(tab4, text="  Инвайтинг  ")
        tab5 = ttk.Frame(nb); nb.add(tab5, text="  База  ")

        self._build_conn_tab(tab1)
        self._build_mon_tab(tab2)
        self._build_mail_tab(tab3)
        self._build_invite_tab(tab4)
        self._build_export_tab(tab5)

        self.status_var = tk.StringVar(value="Не подключено")
        ttk.Label(self.root, textvariable=self.status_var,
                  relief=tk.SUNKEN, anchor=tk.W,
                  padding=(6, 2)).pack(fill=tk.X, padx=8, pady=(0, 4))

    def _build_conn_tab(self, parent):
        f = ttk.LabelFrame(parent, text="Данные Telegram API", padding=14)
        f.pack(fill=tk.X, padx=24, pady=18)

        fields = [
            ("API ID:",                  "api_id_var",   False),
            ("API Hash:",                "api_hash_var", False),
            ("Номер телефона (+7…):",    "phone_var",    False),
        ]
        for row, (label, attr, secret) in enumerate(fields):
            ttk.Label(f, text=label).grid(row=row, column=0, sticky=tk.W, pady=5, padx=(0, 8))
            var = tk.StringVar(); setattr(self, attr, var)
            kw = {"show": "*"} if secret else {}
            entry = ttk.Entry(f, textvariable=var, width=40, **kw)
            entry.grid(row=row, column=1, pady=5, sticky=tk.W)
            _bind_edit_menu(self.root, entry)

        ttk.Label(
            f,
            text="Получить API ID и Hash: https://my.telegram.org  →  API development tools",
            foreground="#0055cc"
        ).grid(row=3, column=0, columnspan=2, pady=(10, 2), sticky=tk.W)

        ttk.Label(
            f,
            text="Вставка: Cmd+V (Mac) или Ctrl+V (Win)  |  ПКМ по полю — Копировать / Вставить",
            foreground="#666666"
        ).grid(row=4, column=0, columnspan=2, pady=(0, 2), sticky=tk.W)

        bf = ttk.Frame(parent); bf.pack(pady=8)
        self.connect_btn = ttk.Button(bf, text="Подключиться", command=self._connect)
        self.connect_btn.pack(side=tk.LEFT, padx=5)
        self.disconnect_btn = ttk.Button(bf, text="Отключиться",
                                          command=self._disconnect, state=tk.DISABLED)
        self.disconnect_btn.pack(side=tk.LEFT, padx=5)

        # Code frame (hidden until needed)
        self._code_frame = ttk.LabelFrame(parent, text="Код подтверждения", padding=10)
        ttk.Label(self._code_frame, text="Код из Telegram / SMS:").grid(row=0, column=0, padx=(0, 8))
        self.code_var = tk.StringVar()
        code_entry = ttk.Entry(self._code_frame, textvariable=self.code_var, width=14)
        code_entry.grid(row=0, column=1)
        _bind_edit_menu(self.root, code_entry)
        ttk.Button(self._code_frame, text="Подтвердить",
                   command=self._submit_code).grid(row=0, column=2, padx=(8, 0))
        self._resend_btn = ttk.Button(self._code_frame, text="Отправить код снова",
                                      command=self._resend_code)
        self._resend_btn.grid(row=1, column=0, columnspan=3, pady=(8, 0), sticky=tk.W)

        # 2FA frame (hidden until needed)
        self._pw_frame = ttk.LabelFrame(parent, text="Двухфакторная аутентификация (2FA)", padding=10)
        ttk.Label(self._pw_frame, text="Пароль 2FA:").grid(row=0, column=0, padx=(0, 8))
        self.password_var = tk.StringVar()
        pw_entry = ttk.Entry(self._pw_frame, textvariable=self.password_var,
                             show="*", width=24)
        pw_entry.grid(row=0, column=1)
        _bind_edit_menu(self.root, pw_entry)
        ttk.Button(self._pw_frame, text="Войти",
                   command=self._submit_password).grid(row=0, column=2, padx=(8, 0))

    def _build_mon_tab(self, parent):
        cf = ttk.LabelFrame(parent, text="Где мониторить", padding=10)
        cf.pack(fill=tk.X, padx=16, pady=(10, 4))
        chat_row = ttk.Frame(cf)
        chat_row.pack(fill=tk.X)
        self.mon_load_chats_btn = ttk.Button(
            chat_row, text="🔄 Обновить чаты",
            command=self._load_chats, state=tk.DISABLED)
        self.mon_load_chats_btn.pack(side=tk.LEFT)
        self.monitor_chat_var = tk.StringVar(value="Все чаты")
        self.monitor_chat_combo = ttk.Combobox(
            chat_row, textvariable=self.monitor_chat_var, width=52, state="readonly")
        self.monitor_chat_combo.pack(side=tk.LEFT, padx=8, fill=tk.X, expand=True)
        ttk.Label(
            cf,
            text="«Все чаты» — мониторинг везде. Или выберите одну беседу / канал.",
            foreground="#666666",
        ).pack(anchor=tk.W, pady=(6, 0))

        kf = ttk.LabelFrame(parent, text="Ключевые слова", padding=10)
        kf.pack(fill=tk.X, padx=16, pady=6)
        ttk.Label(kf, text="Слова через запятую:").pack(side=tk.LEFT)
        self.keywords_var = tk.StringVar()
        kw_entry = ttk.Entry(kf, textvariable=self.keywords_var, width=55)
        kw_entry.pack(side=tk.LEFT, padx=8)
        _bind_edit_menu(self.root, kw_entry)

        hf = ttk.LabelFrame(parent, text="Поиск по истории", padding=10)
        hf.pack(fill=tk.X, padx=16, pady=4)
        self.scan_history_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            hf,
            text="Искать в старых сообщениях (вся доступная история чатов)",
            variable=self.scan_history_var,
        ).pack(anchor=tk.W)
        limit_row = ttk.Frame(hf)
        limit_row.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(limit_row, text="Лимит сообщений на чат (0 = без лимита):").pack(side=tk.LEFT)
        self.history_limit_var = tk.StringVar(value="0")
        limit_entry = ttk.Entry(limit_row, textvariable=self.history_limit_var, width=8)
        limit_entry.pack(side=tk.LEFT, padx=8)
        _bind_edit_menu(self.root, limit_entry)

        bf = ttk.Frame(parent); bf.pack(pady=4)
        self.start_btn = ttk.Button(bf, text="▶  Начать мониторинг",
                                     command=self._start_monitoring, state=tk.DISABLED)
        self.start_btn.pack(side=tk.LEFT, padx=5)
        self.scan_btn = ttk.Button(bf, text="🔍  Сканировать историю",
                                   command=self._start_history_scan, state=tk.DISABLED)
        self.scan_btn.pack(side=tk.LEFT, padx=5)
        self.stop_btn = ttk.Button(bf, text="⏹  Остановить",
                                    command=self._stop_monitoring, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)
        ttk.Button(bf, text="Очистить список",
                   command=self._clear_results).pack(side=tk.LEFT, padx=5)

        self.scan_progress_var = tk.StringVar(value="")
        ttk.Label(parent, textvariable=self.scan_progress_var, foreground="#555555").pack(pady=(0, 2))

        rf = ttk.LabelFrame(parent, text="Результаты (новые + история)", padding=8)
        rf.pack(fill=tk.BOTH, expand=True, padx=16, pady=6)
        self.results_text = scrolledtext.ScrolledText(
            rf, state=tk.DISABLED, wrap=tk.WORD, font=("Consolas", 9))
        self.results_text.pack(fill=tk.BOTH, expand=True)
        _bind_edit_menu(self.root, self.results_text, is_text=True, readonly=True)

        self.matched_count_var = tk.StringVar(value="Найдено уникальных пользователей: 0")
        ttk.Label(parent, textvariable=self.matched_count_var).pack(pady=3)

    def _build_mail_tab(self, parent):
        imp = ttk.LabelFrame(parent, text="Импорт и ручной ввод username", padding=8)
        imp.pack(fill=tk.X, padx=16, pady=(10, 4))
        ttk.Label(
            imp,
            text="Username по одному на строку (можно с @). Формат Excel: «Ключ» → категории → @username + описание.",
            wraplength=880,
        ).pack(anchor=tk.W, pady=(0, 6))
        self.manual_usernames_text = scrolledtext.ScrolledText(imp, height=4, wrap=tk.WORD)
        self.manual_usernames_text.pack(fill=tk.X)
        _bind_edit_menu(self.root, self.manual_usernames_text, is_text=True)

        imp_btns = ttk.Frame(imp)
        imp_btns.pack(fill=tk.X, pady=(6, 0))
        ttk.Button(imp_btns, text="Добавить вручную", command=self._add_manual_usernames).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(imp_btns, text="Импорт Excel (.xlsx)", command=self._import_leads_xlsx).pack(side=tk.LEFT, padx=4)
        ttk.Button(imp_btns, text="Удалить импорт/ручные", command=self._remove_imported_leads).pack(side=tk.LEFT, padx=4)

        lf = ttk.LabelFrame(parent, text="Лиды — выберите кому отправить", padding=8)
        lf.pack(fill=tk.BOTH, expand=True, padx=16, pady=(10, 4))

        filter_row = ttk.Frame(lf)
        filter_row.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(filter_row, text="Фильтр:").pack(side=tk.LEFT)
        self.lead_filter_var = tk.StringVar()
        filter_entry = ttk.Entry(filter_row, textvariable=self.lead_filter_var, width=40)
        filter_entry.pack(side=tk.LEFT, padx=8)
        _bind_edit_menu(self.root, filter_entry)
        self.lead_filter_var.trace_add("write", lambda *_: self._refresh_leads_list())

        sel_row = ttk.Frame(lf)
        sel_row.pack(fill=tk.X, pady=(0, 6))
        ttk.Button(sel_row, text="Выбрать все", command=lambda: self._select_all_leads(True)).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(sel_row, text="Снять выбор", command=lambda: self._select_all_leads(False)).pack(side=tk.LEFT, padx=4)
        ttk.Button(sel_row, text="Инвертировать", command=self._invert_lead_selection).pack(side=tk.LEFT, padx=4)
        ttk.Label(sel_row, text="  (двойной клик — переключить выбор)", foreground="#666666").pack(side=tk.LEFT, padx=8)

        tree_wrap = ttk.Frame(lf)
        tree_wrap.pack(fill=tk.BOTH, expand=True)
        cols = ("name", "username", "chat", "message", "date", "source")
        self.leads_tree = ttk.Treeview(tree_wrap, columns=cols, show="tree headings", height=8)
        self.leads_tree.heading("#0", text="✓")
        self.leads_tree.column("#0", width=32, stretch=False)
        self.leads_tree.heading("name", text="Имя")
        self.leads_tree.heading("username", text="Username")
        self.leads_tree.heading("chat", text="Ключ/Чат")
        self.leads_tree.heading("message", text="Описание")
        self.leads_tree.heading("date", text="Дата")
        self.leads_tree.heading("source", text="Источник")
        self.leads_tree.column("name", width=110)
        self.leads_tree.column("username", width=90)
        self.leads_tree.column("chat", width=90)
        self.leads_tree.column("message", width=180)
        self.leads_tree.column("date", width=90)
        self.leads_tree.column("source", width=80)
        tree_scroll = ttk.Scrollbar(tree_wrap, orient=tk.VERTICAL, command=self.leads_tree.yview)
        self.leads_tree.configure(yscrollcommand=tree_scroll.set)
        self.leads_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.leads_tree.bind("<Double-Button-1>", self._toggle_lead_selection)

        mf = ttk.LabelFrame(parent, text="Текст сообщения для рассылки", padding=10)
        mf.pack(fill=tk.X, padx=16, pady=6)
        self.mail_text = scrolledtext.ScrolledText(mf, height=5, wrap=tk.WORD)
        self.mail_text.pack(fill=tk.X)
        _bind_edit_menu(self.root, self.mail_text, is_text=True)

        of = ttk.Frame(parent); of.pack(fill=tk.X, padx=16, pady=5)
        ttk.Label(of, text="Задержка между сообщениями (сек):").pack(side=tk.LEFT)
        self.delay_var = tk.StringVar(value="3")
        delay_entry = ttk.Entry(of, textvariable=self.delay_var, width=6)
        delay_entry.pack(side=tk.LEFT, padx=6)
        _bind_edit_menu(self.root, delay_entry)

        self.send_btn = ttk.Button(
            parent,
            text="Отправить выбранным (0/0)",
            command=self._start_mailing,
            state=tk.DISABLED
        )
        self.send_btn.pack(pady=6)

        lf = ttk.LabelFrame(parent, text="Лог рассылки", padding=8)
        lf.pack(fill=tk.BOTH, expand=True, padx=16, pady=6)
        self.mail_log = scrolledtext.ScrolledText(
            lf, state=tk.DISABLED, wrap=tk.WORD, font=("Consolas", 9))
        self.mail_log.pack(fill=tk.BOTH, expand=True)
        _bind_edit_menu(self.root, self.mail_log, is_text=True, readonly=True)

        self.mail_progress_var = tk.StringVar()
        ttk.Label(parent, textvariable=self.mail_progress_var).pack(pady=3)

    def _build_invite_tab(self, parent):
        target_f = ttk.LabelFrame(parent, text="Куда приглашать — беседа или канал", padding=8)
        target_f.pack(fill=tk.X, padx=16, pady=(10, 4))

        chat_row = ttk.Frame(target_f)
        chat_row.pack(fill=tk.X)
        self.load_chats_btn = ttk.Button(
            chat_row, text="🔄 Обновить список чатов",
            command=self._load_chats, state=tk.DISABLED)
        self.load_chats_btn.pack(side=tk.LEFT)
        self.invite_chat_var = tk.StringVar()
        self.invite_chat_combo = ttk.Combobox(
            chat_row, textvariable=self.invite_chat_var, width=55, state="readonly")
        self.invite_chat_combo.pack(side=tk.LEFT, padx=8, fill=tk.X, expand=True)

        ctrl_row = ttk.Frame(target_f)
        ctrl_row.pack(fill=tk.X, pady=(8, 0))
        ttk.Label(ctrl_row, text="Задержка между приглашениями (сек):").pack(side=tk.LEFT)
        self.invite_delay_var = tk.StringVar(value="5")
        invite_delay_entry = ttk.Entry(ctrl_row, textvariable=self.invite_delay_var, width=6)
        invite_delay_entry.pack(side=tk.LEFT, padx=6)
        _bind_edit_menu(self.root, invite_delay_entry)
        self.invite_start_btn = ttk.Button(
            ctrl_row, text="▶  Начать инвайтинг (0/0)",
            command=self._start_inviting, state=tk.DISABLED)
        self.invite_start_btn.pack(side=tk.LEFT, padx=8)
        self.invite_stop_btn = ttk.Button(
            ctrl_row, text="⏹  Остановить",
            command=self._stop_inviting, state=tk.DISABLED)
        self.invite_stop_btn.pack(side=tk.LEFT, padx=4)

        imp = ttk.LabelFrame(parent, text="Участники — импорт и ручной ввод", padding=8)
        imp.pack(fill=tk.X, padx=16, pady=4)
        ttk.Label(
            imp,
            text="Username по строке или Excel (формат как в рассылке). Можно взять лидов из парсера.",
            wraplength=880,
        ).pack(anchor=tk.W, pady=(0, 4))
        self.invite_manual_text = scrolledtext.ScrolledText(imp, height=3, wrap=tk.WORD)
        self.invite_manual_text.pack(fill=tk.X)
        _bind_edit_menu(self.root, self.invite_manual_text, is_text=True)

        imp_btns = ttk.Frame(imp)
        imp_btns.pack(fill=tk.X, pady=(6, 0))
        ttk.Button(imp_btns, text="Добавить вручную",
                   command=self._add_invite_manual).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(imp_btns, text="Импорт Excel",
                   command=self._import_invite_xlsx).pack(side=tk.LEFT, padx=4)
        ttk.Button(imp_btns, text="Из выгрузки",
                   command=self._import_invite_from_export).pack(side=tk.LEFT, padx=4)
        ttk.Button(imp_btns, text="Из парсера",
                   command=self._copy_leads_to_invite).pack(side=tk.LEFT, padx=4)
        ttk.Button(imp_btns, text="Очистить список",
                   command=self._clear_invite_users).pack(side=tk.LEFT, padx=4)

        lf = ttk.LabelFrame(parent, text="Список для инвайтинга", padding=8)
        lf.pack(fill=tk.BOTH, expand=True, padx=16, pady=4)

        filter_row = ttk.Frame(lf)
        filter_row.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(filter_row, text="Фильтр:").pack(side=tk.LEFT)
        self.invite_filter_var = tk.StringVar()
        invite_filter_entry = ttk.Entry(filter_row, textvariable=self.invite_filter_var, width=40)
        invite_filter_entry.pack(side=tk.LEFT, padx=8)
        _bind_edit_menu(self.root, invite_filter_entry)
        self.invite_filter_var.trace_add("write", lambda *_: self._refresh_invite_list())

        sel_row = ttk.Frame(lf)
        sel_row.pack(fill=tk.X, pady=(0, 6))
        ttk.Button(sel_row, text="Выбрать все",
                   command=lambda: self._select_all_invite(True)).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(sel_row, text="Снять выбор",
                   command=lambda: self._select_all_invite(False)).pack(side=tk.LEFT, padx=4)
        ttk.Button(sel_row, text="Инвертировать",
                   command=self._invert_invite_selection).pack(side=tk.LEFT, padx=4)
        ttk.Label(sel_row, text="  (двойной клик — переключить)", foreground="#666666").pack(side=tk.LEFT, padx=8)

        tree_wrap = ttk.Frame(lf)
        tree_wrap.pack(fill=tk.BOTH, expand=True)
        cols = ("name", "username", "chat", "message", "source")
        self.invite_tree = ttk.Treeview(tree_wrap, columns=cols, show="tree headings", height=7)
        self.invite_tree.heading("#0", text="✓")
        self.invite_tree.column("#0", width=32, stretch=False)
        self.invite_tree.heading("name", text="Имя")
        self.invite_tree.heading("username", text="Username")
        self.invite_tree.heading("chat", text="Ключ/Чат")
        self.invite_tree.heading("message", text="Описание")
        self.invite_tree.heading("source", text="Источник")
        self.invite_tree.column("name", width=120)
        self.invite_tree.column("username", width=100)
        self.invite_tree.column("chat", width=100)
        self.invite_tree.column("message", width=200)
        self.invite_tree.column("source", width=80)
        inv_scroll = ttk.Scrollbar(tree_wrap, orient=tk.VERTICAL, command=self.invite_tree.yview)
        self.invite_tree.configure(yscrollcommand=inv_scroll.set)
        self.invite_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        inv_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.invite_tree.bind("<Double-Button-1>", self._toggle_invite_selection)

        log_f = ttk.LabelFrame(parent, text="Лог инвайтинга", padding=8)
        log_f.pack(fill=tk.BOTH, expand=True, padx=16, pady=(4, 6))
        self.invite_log = scrolledtext.ScrolledText(
            log_f, state=tk.DISABLED, wrap=tk.WORD, font=("Consolas", 9), height=5)
        self.invite_log.pack(fill=tk.BOTH, expand=True)
        _bind_edit_menu(self.root, self.invite_log, is_text=True, readonly=True)

        self.invite_progress_var = tk.StringVar()
        ttk.Label(parent, textvariable=self.invite_progress_var).pack(pady=2)

    def _build_export_tab(self, parent):
        target_f = ttk.LabelFrame(parent, text="Беседа для выгрузки участников", padding=8)
        target_f.pack(fill=tk.X, padx=16, pady=(10, 4))

        chat_row = ttk.Frame(target_f)
        chat_row.pack(fill=tk.X)
        self.export_load_chats_btn = ttk.Button(
            chat_row, text="🔄 Обновить список чатов",
            command=self._load_chats, state=tk.DISABLED)
        self.export_load_chats_btn.pack(side=tk.LEFT)
        self.export_chat_var = tk.StringVar()
        self.export_chat_combo = ttk.Combobox(
            chat_row, textvariable=self.export_chat_var, width=55, state="readonly")
        self.export_chat_combo.pack(side=tk.LEFT, padx=8, fill=tk.X, expand=True)

        opt_row = ttk.Frame(target_f)
        opt_row.pack(fill=tk.X, pady=(8, 0))
        self.export_with_bio_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            opt_row,
            text="Подтягивать описание профиля (медленнее, но полнее)",
            variable=self.export_with_bio_var,
        ).pack(side=tk.LEFT)

        btn_row = ttk.Frame(target_f)
        btn_row.pack(fill=tk.X, pady=(8, 0))
        self.export_btn = ttk.Button(
            btn_row, text="📥  Выгрузить участников в Excel",
            command=self._start_export, state=tk.DISABLED)
        self.export_btn.pack(side=tk.LEFT)

        info_f = ttk.LabelFrame(parent, text="Формат файла", padding=8)
        info_f.pack(fill=tk.X, padx=16, pady=4)
        ttk.Label(
            info_f,
            text="Excel как во всей программе: строка «Ключ» → название чата → @username + описание.\n"
                 "Файл можно импортировать в «Инвайтинг» → «Из выгрузки» или «Импорт Excel».",
            wraplength=880,
            justify=tk.LEFT,
        ).pack(anchor=tk.W)

        log_f = ttk.LabelFrame(parent, text="Лог выгрузки", padding=8)
        log_f.pack(fill=tk.BOTH, expand=True, padx=16, pady=6)
        self.export_log = scrolledtext.ScrolledText(
            log_f, state=tk.DISABLED, wrap=tk.WORD, font=("Consolas", 9), height=12)
        self.export_log.pack(fill=tk.BOTH, expand=True)
        _bind_edit_menu(self.root, self.export_log, is_text=True, readonly=True)

        self.export_progress_var = tk.StringVar()
        ttk.Label(parent, textvariable=self.export_progress_var).pack(pady=2)

    # ── Async / thread-safe helpers ───────────────────────────────────────────

    def _run_async(self, coro):
        return asyncio.run_coroutine_threadsafe(coro, self.loop)

    def _ui(self, fn=None, *, key: str | None = None):
        """Обновление UI из фонового потока. key — объединять частые вызовы."""
        if key is not None:
            self._ui_coalesce[key] = fn
            if not self._ui_coalesce_pending:
                self._ui_coalesce_pending = True
                self.root.after(16, self._flush_coalesced_ui)
            return
        if fn is not None:
            self._ui_queue.append(fn)
        self._schedule_poll()

    def _schedule_poll(self):
        if not self._poll_scheduled:
            self._poll_scheduled = True
            self.root.after(16, self._poll_ui_queue)

    def _flush_coalesced_ui(self):
        self._ui_coalesce_pending = False
        callbacks = list(self._ui_coalesce.values())
        self._ui_coalesce.clear()
        for fn in callbacks:
            try:
                fn()
            except Exception:
                pass
        if self._ui_queue:
            self._schedule_poll()

    def _poll_ui_queue(self):
        self._poll_scheduled = False
        processed = 0
        while self._ui_queue and processed < 40:
            fn = self._ui_queue.pop(0)
            try:
                fn()
            except Exception:
                pass
            processed += 1
        if self._ui_queue or self._ui_coalesce:
            self._schedule_poll()

    def _queue_result_line(self, line: str):
        self._pending_result_lines.append(line.rstrip("\n"))
        self._ui(self._flush_pending_results, key="flush_results")

    def _flush_pending_results(self):
        if not self._pending_result_lines:
            return
        chunk = self._pending_result_lines[:200]
        self._pending_result_lines = self._pending_result_lines[200:]
        text = "\n".join(chunk) + "\n"
        with open(RESULTS_FILE, "a", encoding="utf-8") as fh:
            fh.write(text)
        self.results_text.config(state=tk.NORMAL)
        self.results_text.insert(tk.END, text)
        self.results_text.see(tk.END)
        self.results_text.config(state=tk.DISABLED)
        if self._pending_result_lines:
            self._ui(key="flush_results")

    def _set_status(self, text: str):
        self.status_var.set(text)

    # ── Config ────────────────────────────────────────────────────────────────

    def _load_config(self):
        if not os.path.exists(CONFIG_FILE):
            return
        try:
            with open(CONFIG_FILE, encoding="utf-8") as fh:
                cfg = json.load(fh)
            self.api_id_var.set(cfg.get("api_id", ""))
            self.api_hash_var.set(cfg.get("api_hash", ""))
            self.phone_var.set(cfg.get("phone", ""))
        except Exception:
            pass

    def _save_config(self):
        cfg = {
            "api_id":   self.api_id_var.get(),
            "api_hash": self.api_hash_var.get(),
            "phone":    self.phone_var.get(),
        }
        with open(CONFIG_FILE, "w", encoding="utf-8") as fh:
            json.dump(cfg, fh, ensure_ascii=False, indent=2)

    # ── Connection ────────────────────────────────────────────────────────────

    def _connect(self):
        api_id_str = self.api_id_var.get().strip()
        api_hash   = self.api_hash_var.get().strip()
        phone      = self.phone_var.get().strip()
        if not all([api_id_str, api_hash, phone]):
            messagebox.showerror("Ошибка", "Заполните все поля!")
            return
        try:
            api_id = int(api_id_str)
        except ValueError:
            messagebox.showerror("Ошибка", "API ID должен быть числом!")
            return
        self._save_config()
        self._phone = phone
        self.connect_btn.config(state=tk.DISABLED)
        self._set_status("Подключение…")
        self._run_async(self._async_connect(api_id, api_hash, phone))

    async def _async_connect(self, api_id, api_hash, phone):
        from telethon import TelegramClient
        from telethon.errors import FloodWaitError
        try:
            self.client = TelegramClient(SESSION_FILE, api_id, api_hash)
            await self.client.connect()
            if not await self.client.is_user_authorized():
                try:
                    sent = await self.client.send_code_request(phone)
                    code_type = type(sent.type).__name__
                    if "App" in code_type:
                        hint = "Код отправлен в приложение Telegram — откройте Telegram на телефоне"
                    else:
                        hint = "Код отправлен по SMS"
                except FloodWaitError as e:
                    msg = f"Слишком много попыток. Подождите {e.seconds} сек. и попробуйте снова."
                    self._ui(lambda m=msg: self._set_status(m))
                    self._ui(lambda m=msg: messagebox.showwarning("FloodWait", m))
                    self._ui(lambda: self.connect_btn.config(state=tk.NORMAL))
                    return
                self._ui(lambda: self._code_frame.pack(fill=tk.X, padx=24, pady=8))
                self._ui(lambda h=hint: self._set_status(h))
            else:
                me = await self.client.get_me()
                self._ui(lambda m=me: self._on_connected(m))
        except Exception as ex:
            self._ui(lambda e=ex: self._set_status(f"Ошибка подключения: {e}"))
            self._ui(lambda: self.connect_btn.config(state=tk.NORMAL))

    def _resend_code(self):
        self._resend_btn.config(state=tk.DISABLED)
        self._set_status("Запрос нового кода…")
        self._run_async(self._async_resend_code())

    async def _async_resend_code(self):
        from telethon.errors import FloodWaitError
        try:
            sent = await self.client.send_code_request(self._phone)
            code_type = type(sent.type).__name__
            if "App" in code_type:
                hint = "Новый код отправлен в приложение Telegram"
            else:
                hint = "Новый код отправлен по SMS"
            self._ui(lambda h=hint: self._set_status(h))
        except FloodWaitError as e:
            msg = f"Слишком много попыток. Подождите {e.seconds} сек."
            self._ui(lambda m=msg: self._set_status(m))
            self._ui(lambda m=msg: messagebox.showwarning("FloodWait", m))
        except Exception as ex:
            self._ui(lambda e=ex: self._set_status(f"Ошибка: {e}"))
        finally:
            self._ui(lambda: self._resend_btn.config(state=tk.NORMAL))

    def _submit_code(self):
        code = self.code_var.get().strip()
        if not code:
            messagebox.showerror("Ошибка", "Введите код!")
            return
        self._run_async(self._async_submit_code(code))

    async def _async_submit_code(self, code):
        from telethon.errors import SessionPasswordNeededError
        try:
            await self.client.sign_in(self._phone, code)
            me = await self.client.get_me()
            self._ui(lambda: self._code_frame.pack_forget())
            self._ui(lambda m=me: self._on_connected(m))
        except SessionPasswordNeededError:
            self._ui(lambda: self._code_frame.pack_forget())
            self._ui(lambda: self._pw_frame.pack(fill=tk.X, padx=24, pady=8))
            self._ui(lambda: self._set_status("Требуется пароль 2FA"))
        except Exception as ex:
            self._ui(lambda e=ex: self._set_status(f"Ошибка кода: {e}"))

    def _submit_password(self):
        pw = self.password_var.get()
        self._run_async(self._async_submit_password(pw))

    async def _async_submit_password(self, pw):
        try:
            await self.client.sign_in(password=pw)
            me = await self.client.get_me()
            self._ui(lambda: self._pw_frame.pack_forget())
            self._ui(lambda m=me: self._on_connected(m))
        except Exception as ex:
            self._ui(lambda e=ex: self._set_status(f"Ошибка 2FA: {e}"))

    def _on_connected(self, me):
        name = f"{me.first_name or ''} {me.last_name or ''}".strip()
        tag  = f"@{me.username}" if me.username else me.phone
        self._set_status(f"Подключено: {name} ({tag})")
        self.connect_btn.config(state=tk.DISABLED)
        self.disconnect_btn.config(state=tk.NORMAL)
        self.start_btn.config(state=tk.NORMAL)
        self.scan_btn.config(state=tk.NORMAL)
        self.send_btn.config(state=tk.NORMAL)
        self.load_chats_btn.config(state=tk.NORMAL)
        self.mon_load_chats_btn.config(state=tk.NORMAL)
        self.export_load_chats_btn.config(state=tk.NORMAL)
        self.invite_start_btn.config(state=tk.NORMAL)
        self.export_btn.config(state=tk.NORMAL)

    def _disconnect(self):
        self._stop_monitoring()
        self._stop_inviting()
        if self.client:
            self._run_async(self.client.disconnect())
        self.client = None
        self._set_status("Отключено")
        self.connect_btn.config(state=tk.NORMAL)
        self.disconnect_btn.config(state=tk.DISABLED)
        self.start_btn.config(state=tk.DISABLED)
        self.scan_btn.config(state=tk.DISABLED)
        self.send_btn.config(state=tk.DISABLED)
        self.load_chats_btn.config(state=tk.DISABLED)
        self.mon_load_chats_btn.config(state=tk.DISABLED)
        self.export_load_chats_btn.config(state=tk.DISABLED)
        self.invite_start_btn.config(state=tk.DISABLED)
        self.export_btn.config(state=tk.DISABLED)

    # ── Monitoring ────────────────────────────────────────────────────────────

    def _parse_keywords(self) -> list | None:
        kws = [" ".join(k.split()).lower() for k in self.keywords_var.get().split(",") if k.strip()]
        if not kws:
            messagebox.showerror("Ошибка", "Введите ключевые слова!")
            return None
        return kws

    def _history_limit(self):
        try:
            val = int(self.history_limit_var.get().strip())
            return None if val <= 0 else val
        except ValueError:
            return None

    def _text_matches(self, raw: str) -> str | None:
        text = " ".join((raw or "").split()).lower()
        if not text or not any(kw in text for kw in self._keywords):
            return None
        return text

    def _format_match_line(self, prefix: str, c_name: str, s_name: str, sender, text: str) -> str:
        uname = f"@{sender.username}" if getattr(sender, "username", None) else "нет username"
        return f"[{prefix}]  Чат: {c_name}  |  От: {s_name}  |  Username: {uname}  |  {text}\n"

    def _register_match(self, sender, chat, text: str, msg_date=None, *, source: str = "live"):
        if not hasattr(sender, "id") or not sender.id:
            return

        s_name = _entity_name(sender)
        c_name = _chat_title(chat)
        if msg_date:
            prefix = msg_date.strftime("%Y-%m-%d %H:%M") if hasattr(msg_date, "strftime") else str(msg_date)
            if source == "history":
                prefix = f"История {prefix}"
        else:
            prefix = datetime.now().strftime("%H:%M:%S")
            if source == "live":
                pass
            elif source == "history":
                prefix = f"История {prefix}"

        entry = {
            "_key":     f"tg_{sender.id}",
            "id":       sender.id,
            "name":     s_name,
            "username": getattr(sender, "username", None),
            "chat":     c_name,
            "message":  text[:300],
            "date":     prefix,
            "selected": True,
            "source":   "парсер",
        }

        existing = next((u for u in self.matched_users if u.get("_key") == entry["_key"]), None)
        if existing:
            existing["chat"] = c_name
            existing["message"] = text[:300]
            existing["date"] = prefix
        else:
            self.matched_users.append(entry)

        line = self._format_match_line(prefix, c_name, s_name, sender, text)
        self._queue_result_line(line)
        self._ui(self._update_count, key="update_count")

    def _get_selected_monitor_chat(self):
        label = self.monitor_chat_var.get().strip()
        if not label or label == "Все чаты":
            return None
        return self._get_chat_entity_by_label(label)

    def _monitor_scope_label(self) -> str:
        label = self.monitor_chat_var.get().strip()
        return label if label else "Все чаты"

    def _validate_monitor_chat(self) -> bool:
        label = self.monitor_chat_var.get().strip()
        if not label or label == "Все чаты":
            return True
        if self._get_chat_entity_by_label(label) is not None:
            return True
        messagebox.showerror(
            "Мониторинг",
            "Выберите чат из списка.\nНажмите «Обновить чаты» после подключения.",
        )
        return False

    def _start_history_scan(self):
        kws = self._parse_keywords()
        if not kws:
            return
        if not self._validate_monitor_chat():
            return
        self._keywords = kws
        self._monitor_chat_entity = self._get_selected_monitor_chat()
        self.is_scanning = True
        self.scan_btn.config(state=tk.DISABLED)
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        scope = self._monitor_scope_label()
        self._set_status(f"Сканирование истории  |  {scope}  |  слова: {', '.join(kws)}")
        self._run_async(self._async_scan_history())

    def _start_monitoring(self):
        kws = self._parse_keywords()
        if not kws:
            return
        if not self._validate_monitor_chat():
            return
        self._keywords = kws
        self._monitor_chat_entity = self._get_selected_monitor_chat()
        self.is_monitoring = True
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        scope = self._monitor_scope_label()
        self._set_status(f"Мониторинг: {scope}  |  слова: {', '.join(kws)}")
        self._run_async(self._async_add_handler())
        if self.scan_history_var.get():
            self.is_scanning = True
            self.scan_btn.config(state=tk.DISABLED)
            self._run_async(self._async_scan_history())

    async def _async_scan_history(self):
        try:
            limit = self._history_limit()
            found = 0

            if self._monitor_chat_entity is not None:
                chats = [self._monitor_chat_entity]
            else:
                dialogs = await self.client.get_dialogs()
                chats = [d.entity for d in dialogs]

            total = len(chats)
            for idx, chat in enumerate(chats, 1):
                if not self.is_scanning:
                    break

                c_name = _chat_title(chat)
                status = f"Сканирование [{idx}/{total}]: {c_name}"
                self._ui(lambda s=status: self.scan_progress_var.set(s), key="scan_progress")
                self._ui(lambda s=status: self._set_status(s), key="status")

                try:
                    async for message in self.client.iter_messages(chat, limit=limit):
                        if not self.is_scanning:
                            break
                        matched = self._text_matches(message.text or message.message or "")
                        if not matched:
                            continue
                        try:
                            sender = await message.get_sender()
                            if sender is None:
                                continue
                            self._register_match(sender, chat, matched, message.date, source="history")
                            found += 1
                        except Exception as ex:
                            self._queue_result_line(f"[Ошибка истории в «{c_name}»: {ex}]")
                except Exception as ex:
                    self._queue_result_line(f"[Не удалось прочитать «{c_name}»: {ex}]")

            done = f"Сканирование завершено. Совпадений: {found}, чатов: {total}"
            self._ui(lambda d=done: self.scan_progress_var.set(d), key="scan_progress")
            if self.is_monitoring:
                self._ui(lambda: self._set_status(
                    f"Мониторинг активен  |  слова: {', '.join(self._keywords)}"
                ), key="status")
            else:
                self._ui(lambda d=done: self._set_status(d), key="status")
        finally:
            self.is_scanning = False
            self._ui(self._on_scan_finished)

    def _on_scan_finished(self):
        if self.client and not self.is_monitoring:
            self.start_btn.config(state=tk.NORMAL)
            self.scan_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)
        elif self.client:
            self.scan_btn.config(state=tk.NORMAL)

    async def _async_add_handler(self):
        from telethon import events

        async def on_message(event):
            if not self.is_monitoring:
                return
            matched = self._text_matches(event.message.text or event.message.message or "")
            if not matched:
                return
            try:
                sender = await event.get_sender()
                chat   = await event.get_chat()
                self._register_match(sender, chat, matched, source="live")
            except Exception as ex:
                self._queue_result_line(f"[Ошибка обработки: {ex}]")

        self._active_handler = on_message
        if self._monitor_chat_entity is not None:
            self.client.add_event_handler(
                on_message, events.NewMessage(chats=[self._monitor_chat_entity]))
        else:
            self.client.add_event_handler(on_message, events.NewMessage)

    def _stop_monitoring(self):
        was_active = self.is_monitoring or self.is_scanning
        self.is_monitoring = False
        self.is_scanning = False
        if self.client and self._active_handler:
            handler = self._active_handler
            self._run_async(self._async_remove_handler(handler))
        self._active_handler = None
        self.stop_btn.config(state=tk.DISABLED)
        if was_active:
            self._save_summary()
        if self.client:
            self.start_btn.config(state=tk.NORMAL)
            self.scan_btn.config(state=tk.NORMAL)
            if was_active:
                self._set_status("Остановлено")
            self.scan_progress_var.set("")

    def _save_summary(self):
        if not self.matched_users:
            return
        ts_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sep    = "=" * 60
        lines  = [
            f"\n{sep}",
            f"СЕССИЯ ЗАВЕРШЕНА: {ts_now}",
            f"Найдено уникальных пользователей: {len(self.matched_users)}",
            sep,
        ]
        for i, u in enumerate(self.matched_users, 1):
            uname = f"@{u['username']}" if u["username"] else "нет username"
            msg = (u.get("message") or "")[:60]
            lines.append(f"  {i}. {u['name']}  |  {uname}  |  Чат: {u['chat']}  |  {msg}")
        lines.append(sep + "\n")
        with open(RESULTS_FILE, "a", encoding="utf-8") as fh:
            fh.write("\n".join(lines))

    async def _async_remove_handler(self, handler):
        self.client.remove_event_handler(handler)

    def _add_result(self, text: str):
        self.results_text.config(state=tk.NORMAL)
        self.results_text.insert(tk.END, text)
        self.results_text.see(tk.END)
        self.results_text.config(state=tk.DISABLED)

    def _update_count(self):
        self.matched_count_var.set(
            f"Найдено уникальных пользователей: {len(self.matched_users)}"
        )
        self._refresh_leads_list()

    def _clear_results(self):
        self.results_text.config(state=tk.NORMAL)
        self.results_text.delete(1.0, tk.END)
        self.results_text.config(state=tk.DISABLED)
        self.matched_users.clear()
        self._update_count()

    def _lead_source_label(self, user: dict) -> str:
        source = user.get("source", "парсер")
        labels = {"import": "Excel", "manual": "Вручную", "парсер": "Парсер"}
        return labels.get(source, source)

    def _lead_key(self, user: dict) -> str:
        if user.get("_key"):
            return user["_key"]
        if user.get("id"):
            user["_key"] = f"tg_{user['id']}"
        elif user.get("username"):
            user["_key"] = f"u_{_normalize_username(user['username']).lower()}"
        else:
            user["_key"] = f"row_{id(user)}"
        return user["_key"]

    def _make_lead_entry(
        self,
        *,
        username: str,
        name: str = "",
        chat: str = "",
        message: str = "",
        source: str = "manual",
        user_id=None,
    ) -> dict | None:
        username = _normalize_username(username)
        if not username:
            return None
        return {
            "_key": f"tg_{user_id}" if user_id else f"u_{username.lower()}",
            "id": user_id,
            "name": name or f"@{username}",
            "username": username,
            "chat": chat or ("Ручной ввод" if source == "manual" else "Импорт Excel"),
            "message": (message or "")[:300],
            "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "selected": True,
            "source": source,
        }

    def _parse_manual_usernames(self, text: str) -> list[str]:
        found: list[str] = []
        seen: set[str] = set()
        for line in text.splitlines():
            chunk = line.replace(";", ",")
            for part in chunk.split(","):
                username = _normalize_username(part)
                if not username:
                    continue
                key = username.lower()
                if key in seen:
                    continue
                seen.add(key)
                found.append(username)
        return found

    def _upsert_leads(self, leads: list[dict]) -> tuple[int, int]:
        added = 0
        updated = 0
        for raw in leads:
            entry = self._make_lead_entry(
                username=raw.get("username", ""),
                name=raw.get("name", ""),
                chat=raw.get("chat", ""),
                message=raw.get("message", ""),
                source=raw.get("source", "manual"),
                user_id=raw.get("id"),
            )
            if not entry:
                continue
            existing = next((u for u in self.matched_users if u.get("_key") == entry["_key"]), None)
            if existing:
                existing.update({
                    "name": entry["name"],
                    "chat": entry["chat"],
                    "message": entry["message"],
                    "date": entry["date"],
                    "source": entry["source"],
                    "username": entry["username"],
                })
                updated += 1
            else:
                self.matched_users.append(entry)
                added += 1
        self._update_count()
        return added, updated

    def _add_manual_usernames(self):
        text = self.manual_usernames_text.get(1.0, tk.END)
        usernames = self._parse_manual_usernames(text)
        if not usernames:
            messagebox.showinfo("Ручной ввод", "Введите хотя бы один username.")
            return
        leads = [{"username": u, "source": "manual"} for u in usernames]
        added, updated = self._upsert_leads(leads)
        self.manual_usernames_text.delete(1.0, tk.END)
        messagebox.showinfo(
            "Ручной ввод",
            f"Добавлено: {added}\nОбновлено: {updated}\nВсего в списке: {len(self.matched_users)}"
        )

    def _import_leads_xlsx(self):
        path = filedialog.askopenfilename(
            title="Импорт базы Excel",
            filetypes=[("Excel", "*.xlsx"), ("Все файлы", "*.*")],
        )
        if not path:
            return
        try:
            leads = parse_leads_xlsx(path)
        except Exception as ex:
            messagebox.showerror("Импорт Excel", f"Не удалось прочитать файл:\n{ex}")
            return
        if not leads:
            messagebox.showinfo("Импорт Excel", "В файле не найдено username.")
            return
        added, updated = self._upsert_leads(leads)
        messagebox.showinfo(
            "Импорт Excel",
            f"Файл: {os.path.basename(path)}\n"
            f"Найдено username: {len(leads)}\n"
            f"Добавлено: {added}\n"
            f"Обновлено: {updated}\n"
            f"Всего в списке: {len(self.matched_users)}"
        )

    def _remove_imported_leads(self):
        before = len(self.matched_users)
        self.matched_users = [
            u for u in self.matched_users
            if u.get("source") not in {"manual", "import"}
        ]
        removed = before - len(self.matched_users)
        self._update_count()
        messagebox.showinfo("Очистка", f"Удалено импортированных/ручных: {removed}")

    def _lead_filter_text(self) -> str:
        return self.lead_filter_var.get().strip().lower() if hasattr(self, "lead_filter_var") else ""

    def _lead_matches_filter(self, user: dict) -> bool:
        needle = self._lead_filter_text()
        if not needle:
            return True
        haystack = " ".join([
            user.get("name", ""),
            user.get("username") or "",
            user.get("chat", ""),
            user.get("message", ""),
            user.get("date", ""),
            self._lead_source_label(user),
        ]).lower()
        return needle in haystack

    def _refresh_leads_list(self):
        if not hasattr(self, "leads_tree"):
            return
        for item in self.leads_tree.get_children():
            self.leads_tree.delete(item)
        for user in self.matched_users:
            if not self._lead_matches_filter(user):
                continue
            mark = "☑" if user.get("selected", True) else "☐"
            uname = f"@{user['username']}" if user.get("username") else "—"
            msg = (user.get("message") or "")[:80]
            self.leads_tree.insert(
                "", tk.END,
                iid=self._lead_key(user),
                text=mark,
                values=(
                    user["name"],
                    uname,
                    user.get("chat", ""),
                    msg,
                    user.get("date", ""),
                    self._lead_source_label(user),
                ),
            )
        self._update_mail_selection_count()

    def _toggle_lead_selection(self, event):
        region = self.leads_tree.identify_region(event.x, event.y)
        if region not in ("tree", "cell"):
            return
        item = self.leads_tree.identify_row(event.y)
        if not item:
            return
        for user in self.matched_users:
            if self._lead_key(user) == item:
                user["selected"] = not user.get("selected", True)
                break
        self._refresh_leads_list()

    def _select_all_leads(self, selected: bool):
        needle = self._lead_filter_text()
        for user in self.matched_users:
            if needle and not self._lead_matches_filter(user):
                continue
            user["selected"] = selected
        self._refresh_leads_list()

    def _invert_lead_selection(self):
        needle = self._lead_filter_text()
        for user in self.matched_users:
            if needle and not self._lead_matches_filter(user):
                continue
            user["selected"] = not user.get("selected", True)
        self._refresh_leads_list()

    def _get_selected_users(self) -> list:
        return [u for u in self.matched_users if u.get("selected", True)]

    def _update_mail_selection_count(self):
        if not hasattr(self, "send_btn"):
            return
        selected = len(self._get_selected_users())
        total = len(self.matched_users)
        self.send_btn.config(text=f"Отправить выбранным ({selected}/{total})")

    # ── Mailing ───────────────────────────────────────────────────────────────

    def _start_mailing(self):
        recipients = self._get_selected_users()
        if not recipients:
            messagebox.showinfo(
                "Рассылка",
                "Никто не выбран. Отметьте лидов на вкладке «Рассылка» или запустите поиск."
            )
            return
        msg = self.mail_text.get(1.0, tk.END).strip()
        if not msg:
            messagebox.showerror("Ошибка", "Введите текст сообщения!")
            return
        try:
            delay = max(0.5, float(self.delay_var.get()))
        except ValueError:
            delay = 3.0
        preview = msg[:80] + ("…" if len(msg) > 80 else "")
        names_preview = "\n".join(
            f"  • {u['name']} ({u.get('chat', '')})" for u in recipients[:8]
        )
        if len(recipients) > 8:
            names_preview += f"\n  … и ещё {len(recipients) - 8}"
        if not messagebox.askyesno(
            "Подтверждение",
            f"Отправить сообщение {len(recipients)} выбранным пользователям?\n\n"
            f"«{preview}»\n\n{names_preview}"
        ):
            return
        self.send_btn.config(state=tk.DISABLED)
        self._run_async(self._async_mailing(msg, delay, recipients))

    async def _resolve_recipient(self, user: dict):
        if user.get("id"):
            return user["id"]
        username = _normalize_username(user.get("username", ""))
        if not username:
            raise ValueError("Нет username для отправки")
        entity = await self.client.get_entity(username)
        user["id"] = entity.id
        user["_key"] = f"tg_{entity.id}"
        if not user.get("name") or user["name"] == f"@{username}":
            user["name"] = _entity_name(entity)
        return entity

    async def _async_mailing(self, message: str, delay: float, recipients: list):
        import asyncio as aio
        total   = len(recipients)
        success = 0
        for i, user in enumerate(recipients, 1):
            label = user.get("name") or (f"@{user.get('username')}" if user.get("username") else "?")
            try:
                target = await self._resolve_recipient(user)
                await self.client.send_message(target, message)
                log = f"[{i}/{total}] ✓  {label}\n"
                success += 1
            except Exception as ex:
                log = f"[{i}/{total}] ✗  {label}: {ex}\n"
            self._ui(lambda l=log: self._add_mail_log(l))
            self._ui(lambda v=f"Прогресс: {i}/{total}": self.mail_progress_var.set(v))
            if i < total:
                await aio.sleep(delay)
        summary = f"\n--- Рассылка завершена. Успешно: {success}/{total} ---\n"
        self._ui(lambda s=summary: self._add_mail_log(s))
        self._ui(self._refresh_leads_list)
        self._ui(lambda: self.send_btn.config(state=tk.NORMAL))

    def _add_mail_log(self, text: str):
        self.mail_log.config(state=tk.NORMAL)
        self.mail_log.insert(tk.END, text)
        self.mail_log.see(tk.END)
        self.mail_log.config(state=tk.DISABLED)

    # ── Inviting ──────────────────────────────────────────────────────────────

    def _invite_source_label(self, user: dict) -> str:
        labels = {"import": "Excel", "manual": "Вручную", "parser": "Парсер", "export": "Выгрузка", "парсер": "Парсер"}
        return labels.get(user.get("source", ""), user.get("source", ""))

    def _upsert_invite_users(self, leads: list[dict]) -> tuple[int, int]:
        added = updated = 0
        for raw in leads:
            entry = self._make_lead_entry(
                username=raw.get("username", ""),
                name=raw.get("name", ""),
                chat=raw.get("chat", ""),
                message=raw.get("message", ""),
                source=raw.get("source", "manual"),
                user_id=raw.get("id"),
            )
            if not entry:
                continue
            existing = next((u for u in self.invite_users if u.get("_key") == entry["_key"]), None)
            if existing:
                existing.update({
                    "name": entry["name"], "chat": entry["chat"],
                    "message": entry["message"], "date": entry["date"],
                    "source": entry["source"], "username": entry["username"],
                })
                updated += 1
            else:
                self.invite_users.append(entry)
                added += 1
        self._refresh_invite_list()
        return added, updated

    def _invite_filter_text(self) -> str:
        return self.invite_filter_var.get().strip().lower() if hasattr(self, "invite_filter_var") else ""

    def _invite_matches_filter(self, user: dict) -> bool:
        needle = self._invite_filter_text()
        if not needle:
            return True
        haystack = " ".join([
            user.get("name", ""), user.get("username") or "",
            user.get("chat", ""), user.get("message", ""),
            self._invite_source_label(user),
        ]).lower()
        return needle in haystack

    def _get_selected_invite_users(self) -> list:
        return [u for u in self.invite_users if u.get("selected", True)]

    def _refresh_invite_list(self):
        if not hasattr(self, "invite_tree"):
            return
        for item in self.invite_tree.get_children():
            self.invite_tree.delete(item)
        for user in self.invite_users:
            if not self._invite_matches_filter(user):
                continue
            mark = "☑" if user.get("selected", True) else "☐"
            uname = f"@{user['username']}" if user.get("username") else "—"
            msg = (user.get("message") or "")[:80]
            self.invite_tree.insert(
                "", tk.END, iid=self._lead_key(user), text=mark,
                values=(user["name"], uname, user.get("chat", ""), msg,
                        self._invite_source_label(user)),
            )
        self._update_invite_count()

    def _update_invite_count(self):
        if not hasattr(self, "invite_start_btn"):
            return
        sel = len(self._get_selected_invite_users())
        total = len(self.invite_users)
        self.invite_start_btn.config(text=f"▶  Начать инвайтинг ({sel}/{total})")

    def _toggle_invite_selection(self, event):
        region = self.invite_tree.identify_region(event.x, event.y)
        if region not in ("tree", "cell"):
            return
        item = self.invite_tree.identify_row(event.y)
        if not item:
            return
        for user in self.invite_users:
            if self._lead_key(user) == item:
                user["selected"] = not user.get("selected", True)
                break
        self._refresh_invite_list()

    def _select_all_invite(self, selected: bool):
        needle = self._invite_filter_text()
        for user in self.invite_users:
            if needle and not self._invite_matches_filter(user):
                continue
            user["selected"] = selected
        self._refresh_invite_list()

    def _invert_invite_selection(self):
        needle = self._invite_filter_text()
        for user in self.invite_users:
            if needle and not self._invite_matches_filter(user):
                continue
            user["selected"] = not user.get("selected", True)
        self._refresh_invite_list()

    def _add_invite_manual(self):
        text = self.invite_manual_text.get(1.0, tk.END)
        usernames = self._parse_manual_usernames(text)
        if not usernames:
            messagebox.showinfo("Инвайтинг", "Введите хотя бы один username.")
            return
        added, updated = self._upsert_invite_users(
            [{"username": u, "source": "manual"} for u in usernames])
        self.invite_manual_text.delete(1.0, tk.END)
        messagebox.showinfo("Инвайтинг", f"Добавлено: {added}\nОбновлено: {updated}")

    def _import_invite_xlsx(self):
        path = filedialog.askopenfilename(
            title="Импорт Excel для инвайтинга",
            filetypes=[("Excel", "*.xlsx"), ("Все файлы", "*.*")],
        )
        if not path:
            return
        try:
            leads = parse_leads_xlsx(path)
        except Exception as ex:
            messagebox.showerror("Импорт Excel", str(ex))
            return
        if not leads:
            messagebox.showinfo("Импорт Excel", "Username не найдены.")
            return
        added, updated = self._upsert_invite_users(leads)
        messagebox.showinfo(
            "Импорт Excel",
            f"Найдено: {len(leads)}\nДобавлено: {added}\nОбновлено: {updated}",
        )

    def _import_invite_from_export(self):
        initialdir = os.path.dirname(self._last_export_path) if self._last_export_path else _APP_DIR
        path = filedialog.askopenfilename(
            title="Импорт из выгрузки (База)",
            initialdir=initialdir,
            filetypes=[("Excel", "*.xlsx"), ("Все файлы", "*.*")],
        )
        if not path:
            return
        try:
            leads = parse_leads_xlsx(path)
        except Exception as ex:
            messagebox.showerror("Импорт", str(ex))
            return
        if not leads:
            messagebox.showinfo("Импорт", "Username не найдены.")
            return
        for lead in leads:
            lead["source"] = "export"
        added, updated = self._upsert_invite_users(leads)
        messagebox.showinfo(
            "Импорт из выгрузки",
            f"Файл: {os.path.basename(path)}\n"
            f"Найдено: {len(leads)}\nДобавлено: {added}\nОбновлено: {updated}",
        )

    def _copy_leads_to_invite(self):
        source = self._get_selected_users() if self.matched_users else []
        if not source:
            source = self.matched_users
        if not source:
            messagebox.showinfo("Инвайтинг", "Список парсера пуст. Сначала найдите лидов.")
            return
        leads = [{
            "username": u.get("username", ""),
            "name": u.get("name", ""),
            "chat": u.get("chat", ""),
            "message": u.get("message", ""),
            "id": u.get("id"),
            "source": "parser",
        } for u in source if u.get("username") or u.get("id")]
        added, updated = self._upsert_invite_users(leads)
        messagebox.showinfo("Инвайтинг", f"Из парсера: +{added}, обновлено {updated}")

    def _clear_invite_users(self):
        self.invite_users.clear()
        self._refresh_invite_list()

    def _load_chats(self):
        if not self.client:
            messagebox.showerror("Ошибка", "Сначала подключитесь к Telegram.")
            return
        self.load_chats_btn.config(state=tk.DISABLED)
        self.mon_load_chats_btn.config(state=tk.DISABLED)
        self.export_load_chats_btn.config(state=tk.DISABLED)
        self._set_status("Загрузка списка чатов…")
        self._run_async(self._async_load_chats())

    async def _async_load_chats(self):
        try:
            dialogs = await self.client.get_dialogs()
            labels: list[str] = []
            self._invite_chat_entities.clear()
            self._invite_label_to_id.clear()
            for d in dialogs:
                if not (d.is_group or d.is_channel):
                    continue
                entity = d.entity
                if getattr(entity, "broadcast", False) and not getattr(entity, "megagroup", False):
                    kind = "канал"
                elif d.is_channel:
                    kind = "супергруппа"
                else:
                    kind = "группа"
                title = _chat_title(entity)
                label = f"[{kind}] {title}"
                base = label
                n = 2
                while label in self._invite_label_to_id:
                    label = f"{base} ({n})"
                    n += 1
                labels.append(label)
                eid = str(entity.id)
                self._invite_label_to_id[label] = eid
                self._invite_chat_entities[eid] = entity

            def update_ui():
                mon_values = ["Все чаты"] + labels
                self.monitor_chat_combo["values"] = mon_values
                cur = self.monitor_chat_var.get().strip()
                if cur not in mon_values:
                    self.monitor_chat_var.set("Все чаты")
                self.invite_chat_combo["values"] = labels
                self.export_chat_combo["values"] = labels
                if labels:
                    if not self.invite_chat_var.get():
                        self.invite_chat_var.set(labels[0])
                    if not self.export_chat_var.get():
                        self.export_chat_var.set(labels[0])
                self._set_status(f"Загружено чатов: {len(labels)}")

            self._ui(update_ui)
        except Exception as ex:
            self._ui(lambda e=ex: messagebox.showerror("Ошибка", f"Не удалось загрузить чаты:\n{e}"))
            self._ui(lambda: self._set_status("Ошибка загрузки чатов"))
        finally:
            self._ui(lambda: self.load_chats_btn.config(state=tk.NORMAL))
            self._ui(lambda: self.mon_load_chats_btn.config(state=tk.NORMAL))
            self._ui(lambda: self.export_load_chats_btn.config(state=tk.NORMAL))

    def _get_chat_entity_by_label(self, label: str):
        if not label or label not in self._invite_label_to_id:
            return None
        eid = self._invite_label_to_id[label]
        return self._invite_chat_entities.get(eid)

    def _get_selected_export_chat(self):
        return self._get_chat_entity_by_label(self.export_chat_var.get().strip())

    def _get_selected_invite_chat(self):
        return self._get_chat_entity_by_label(self.invite_chat_var.get().strip())

    def _start_inviting(self):
        target = self._get_selected_invite_chat()
        if target is None:
            messagebox.showerror("Инвайтинг", "Выберите беседу или канал.\nНажмите «Обновить список чатов».")
            return
        users = self._get_selected_invite_users()
        if not users:
            messagebox.showinfo("Инвайтинг", "Отметьте участников для приглашения.")
            return
        try:
            delay = max(1.0, float(self.invite_delay_var.get()))
        except ValueError:
            delay = 5.0
        chat_name = self.invite_chat_var.get()
        preview = "\n".join(
            f"  • {u.get('name', '?')}" for u in users[:8]
        )
        if len(users) > 8:
            preview += f"\n  … и ещё {len(users) - 8}"
        if not messagebox.askyesno(
            "Подтверждение",
            f"Пригласить {len(users)} человек в:\n{chat_name}\n\n{preview}",
        ):
            return
        self._invite_stop = False
        self.is_inviting = True
        self.invite_start_btn.config(state=tk.DISABLED)
        self.invite_stop_btn.config(state=tk.NORMAL)
        self._run_async(self._async_inviting(target, users, delay, chat_name))

    def _stop_inviting(self):
        self._invite_stop = True
        self.invite_stop_btn.config(state=tk.DISABLED)

    def _add_invite_log(self, text: str):
        self.invite_log.config(state=tk.NORMAL)
        self.invite_log.insert(tk.END, text)
        self.invite_log.see(tk.END)
        self.invite_log.config(state=tk.DISABLED)

    async def _async_invite_user(self, target, user_entity):
        from telethon.tl.types import Channel, Chat
        from telethon.tl.functions.channels import InviteToChannelRequest
        from telethon.tl.functions.messages import AddChatUserRequest

        if isinstance(target, Channel):
            await self.client(InviteToChannelRequest(target, [user_entity]))
        elif isinstance(target, Chat):
            await self.client(AddChatUserRequest(target.id, user_entity, fwd_limit=0))
        else:
            await self.client(InviteToChannelRequest(target, [user_entity]))

    async def _async_inviting(self, target, users: list, delay: float, chat_name: str):
        import asyncio as aio
        from telethon.errors import (
            ChatAdminRequiredError,
            FloodWaitError,
            UserAlreadyParticipantError,
            UserNotMutualContactError,
            UserPrivacyRestrictedError,
        )

        total = len(users)
        success = 0
        try:
            for i, user in enumerate(users, 1):
                if self._invite_stop:
                    self._ui(lambda: self._add_invite_log("\n--- Остановлено пользователем ---\n"))
                    break

                label = user.get("name") or f"@{user.get('username', '?')}"
                try:
                    target_user = await self._resolve_recipient(user)
                    await self._async_invite_user(target, target_user)
                    log = f"[{i}/{total}] ✓  {label}\n"
                    success += 1
                except UserAlreadyParticipantError:
                    log = f"[{i}/{total}] ○  {label} — уже в чате\n"
                    success += 1
                except FloodWaitError as e:
                    log = f"[{i}/{total}] ⏳ FloodWait {e.seconds} сек.\n"
                    self._ui(lambda l=log: self._add_invite_log(l))
                    await aio.sleep(e.seconds)
                    try:
                        target_user = await self._resolve_recipient(user)
                        await self._async_invite_user(target, target_user)
                        log = f"[{i}/{total}] ✓  {label} (повтор)\n"
                        success += 1
                    except Exception as ex2:
                        log = f"[{i}/{total}] ✗  {label}: {ex2}\n"
                except (UserPrivacyRestrictedError, UserNotMutualContactError) as ex:
                    log = f"[{i}/{total}] ✗  {label}: приватность / не контакт\n"
                except ChatAdminRequiredError:
                    log = f"[{i}/{total}] ✗  {label}: нет прав админа для инвайта\n"
                    self._ui(lambda l=log: self._add_invite_log(l))
                    break
                except Exception as ex:
                    log = f"[{i}/{total}] ✗  {label}: {ex}\n"

                self._ui(lambda l=log: self._add_invite_log(l))
                self._ui(lambda v=f"Прогресс: {i}/{total}": self.invite_progress_var.set(v))
                if i < total and not self._invite_stop:
                    await aio.sleep(delay)
        finally:
            summary = f"\n--- Инвайтинг завершён: {success}/{total} → {chat_name} ---\n"
            self._ui(lambda s=summary: self._add_invite_log(s))
            self.is_inviting = False
            self._invite_stop = False
            self._ui(lambda: self.invite_start_btn.config(state=tk.NORMAL))
            self._ui(lambda: self.invite_stop_btn.config(state=tk.DISABLED))
            self._ui(lambda: self.invite_progress_var.set(""))
            self._ui(self._refresh_invite_list)


    # ── Export (База) ─────────────────────────────────────────────────────────

    def _add_export_log(self, text: str):
        self.export_log.config(state=tk.NORMAL)
        self.export_log.insert(tk.END, text if text.endswith("\n") else text + "\n")
        self.export_log.see(tk.END)
        self.export_log.config(state=tk.DISABLED)

    def _start_export(self):
        target = self._get_selected_export_chat()
        if target is None:
            messagebox.showerror("База", "Выберите беседу.\nНажмите «Обновить список чатов».")
            return
        chat_label = self.export_chat_var.get().strip()
        category = chat_label.split("] ", 1)[-1] if "] " in chat_label else chat_label
        safe_name = re.sub(r'[\\/:*?"<>|]', "_", category)[:60]
        default_name = f"база_{safe_name}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Сохранить выгрузку Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            initialdir=_APP_DIR,
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return
        self.is_exporting = True
        self.export_btn.config(state=tk.DISABLED)
        self.export_log.config(state=tk.NORMAL)
        self.export_log.delete(1.0, tk.END)
        self.export_log.config(state=tk.DISABLED)
        with_bio = self.export_with_bio_var.get()
        self._run_async(self._async_export_members(target, category, path, with_bio))

    async def _async_export_members(self, target, category: str, save_path: str, with_bio: bool):
        import asyncio as aio
        from telethon.tl.types import User
        from telethon.tl.functions.users import GetFullUserRequest

        members: list[dict] = []
        skipped = 0
        idx = 0
        try:
            self._ui(lambda: self._add_export_log(f"Выгрузка из: {category}"))
            async for user in self.client.iter_participants(target):
                if not isinstance(user, User) or user.bot:
                    continue
                username = user.username
                if not username:
                    skipped += 1
                    continue

                name = entity_name(user)
                parts = [name]
                if with_bio:
                    try:
                        full = await self.client(GetFullUserRequest(user))
                        if full.full_user.about:
                            parts.append(full.full_user.about.strip())
                        await aio.sleep(0.12)
                    except Exception:
                        pass

                members.append({
                    "username": username,
                    "message": " | ".join(parts)[:500],
                })
                idx += 1
                if idx % 25 == 0:
                    self._ui(
                        lambda c=idx: self.export_progress_var.set(f"Собрано: {c}"),
                        key="export_progress",
                    )
                    self._ui(lambda c=idx: self._add_export_log(f"… обработано {c}"))

            def flush_log():
                self._add_export_log(f"Всего с username: {len(members)}, без username: {skipped}")

            self._ui(flush_log)
            total = write_leads_xlsx(save_path, [(category, members)])
            self._last_export_path = save_path

            def done_ui():
                self._add_export_log(f"✓ Сохранено: {save_path}")
                self._add_export_log(f"✓ Записей в Excel: {total}")
                self.export_progress_var.set(f"Готово: {total} участников")
                self._set_status(f"Выгрузка завершена: {total} → {os.path.basename(save_path)}")
                messagebox.showinfo(
                    "База",
                    f"Выгружено: {total} участников\n"
                    f"Без username (пропущено): {skipped}\n\n{save_path}",
                )

            self._ui(done_ui)
        except Exception as ex:
            self._ui(lambda e=ex: messagebox.showerror("Ошибка выгрузки", str(e)))
            self._ui(lambda e=ex: self._add_export_log(f"✗ Ошибка: {e}"))
        finally:
            self.is_exporting = False
            self._ui(lambda: self.export_btn.config(state=tk.NORMAL))


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
